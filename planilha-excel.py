import openpyxl

#Criar Planilha

book=openpyxl.Workbook()
#Como visualizar páginas existentes
print(book.sheetnames)
#Como criar uma página
book.create_sheet('Frutas')
#Como selecionar uma páginas
frutas_page = book['Frutas']
frutas_page.append(['Item','Quantidade','Preço'])
frutas_page.append(['Banana','5','R$3,90'])
frutas_page.append(['Fruta','2','R$15,90'])
frutas_page.append(['Fruta','10','R$30,90'])
frutas_page.append(['Fruta','2','R$50,50'])
#Salvar a planilha
book.save('Planilha de Compras.xlsx')